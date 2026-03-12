"""Gmail router — server-side device flow (gmail.send) + email sending.

Uses the same GDRIVE_CLIENT_ID / GDRIVE_CLIENT_SECRET Desktop-app credentials.
Per-user Gmail tokens are stored in the gmail_tokens table.
Only the gmail.send scope is requested — inbox is never read.
"""
import base64
import email.mime.application
import email.mime.multipart
import email.mime.text
import io
import re
import time
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..auth.session import get_current_claims
from ..config import settings
from ..database import get_db
from ..models.audit_run import AuditRun
from ..models.gmail_token import GmailToken

router = APIRouter(prefix="/api/gmail", tags=["gmail"])

_GOOGLE_DEVICE_URL = "https://oauth2.googleapis.com/device/code"
_GOOGLE_TOKEN_URL  = "https://oauth2.googleapis.com/token"
_GMAIL_SEND_URL    = "https://gmail.googleapis.com/gmail/v1/users/me/messages/send"
_DEVICE_GRANT      = "urn:ietf:params:oauth:grant-type:device_code"
_GMAIL_SCOPE       = "https://www.googleapis.com/auth/gmail.send"

_MAX_POLL_ATTEMPTS = 20
_EMAIL_RE          = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _require_credentials() -> tuple[str, str]:
    cid = settings.gdrive_client_id.strip()   # reuses same Desktop-app client
    sec = settings.gdrive_client_secret.strip()
    if not cid or not sec:
        raise HTTPException(
            503,
            "Gmail integration is not configured on this server. "
            "Contact the administrator.",
        )
    return cid, sec


async def _get_gmail_token(db: AsyncSession, user_id: str) -> Optional[str]:
    result = await db.execute(
        select(GmailToken).where(GmailToken.user_id == user_id)
    )
    row = result.scalar_one_or_none()
    if not row:
        return None
    if time.time() < row.expires_at - 60:
        return row.access_token
    if not row.refresh_token:
        return None

    cid, sec = _require_credentials()
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            _GOOGLE_TOKEN_URL,
            data={
                "grant_type":    "refresh_token",
                "refresh_token": row.refresh_token,
                "client_id":     cid,
                "client_secret": sec,
            },
        )
    if resp.status_code != 200:
        return None
    tok = resp.json()
    row.access_token = tok["access_token"]
    row.expires_at   = time.time() + tok.get("expires_in", 3600)
    await db.commit()
    return row.access_token


async def _async_sleep(seconds: float) -> None:
    import asyncio
    await asyncio.sleep(seconds)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/authorize/start", summary="Start Gmail send-only device flow")
async def gmail_authorize_start(
    claims: dict = Depends(get_current_claims),
) -> dict:
    """Start the Device Authorization flow for Gmail send-only access.

    Returns ``user_code`` and ``verification_url``. The user opens the URL,
    enters the code, and approves send-only Gmail access in their browser.
    Then call ``/api/gmail/authorize/complete``.

    Only ``gmail.send`` scope is requested — the inbox is never read.
    OAuth credentials never leave the server.
    """
    cid, _ = _require_credentials()
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            _GOOGLE_DEVICE_URL,
            data={"client_id": cid, "scope": _GMAIL_SCOPE},
        )
    if resp.status_code != 200:
        raise HTTPException(502, f"Failed to start Gmail authorization: {resp.text}")
    data = resp.json()
    return {
        "user_code":        data["user_code"],
        "verification_url": data["verification_url"],
        "device_code":      data["device_code"],
        "interval":         data.get("interval", 5),
        "expires_in":       data.get("expires_in", 1800),
        "scope":            "gmail.send — outbound only, inbox never read",
    }


class GmailCompleteRequest(BaseModel):
    device_code: str


@router.post("/authorize/complete", summary="Complete Gmail device flow")
async def gmail_authorize_complete(
    body:   GmailCompleteRequest,
    claims: dict         = Depends(get_current_claims),
    db:     AsyncSession = Depends(get_db),
) -> dict:
    """Poll Google for the Gmail token after the user approved in browser."""
    cid, sec = _require_credentials()
    user_id  = claims["user_id"]

    for _ in range(_MAX_POLL_ATTEMPTS):
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(
                _GOOGLE_TOKEN_URL,
                data={
                    "grant_type":    _DEVICE_GRANT,
                    "device_code":   body.device_code,
                    "client_id":     cid,
                    "client_secret": sec,
                },
            )
        if resp.status_code == 200:
            tok = resp.json()
            result = await db.execute(
                select(GmailToken).where(GmailToken.user_id == user_id)
            )
            row = result.scalar_one_or_none()
            if row is None:
                row = GmailToken(user_id=user_id)
                db.add(row)
            row.access_token  = tok["access_token"]
            row.refresh_token = tok.get("refresh_token", "")
            row.expires_at    = time.time() + tok.get("expires_in", 3600)
            await db.commit()
            return {"ok": True, "message": "Gmail send-only access granted."}

        err = resp.json().get("error", "")
        if err == "authorization_pending":
            await _async_sleep(5)
            continue
        if err == "slow_down":
            await _async_sleep(10)
            continue
        raise HTTPException(400, f"Gmail authorization failed: {err}")

    raise HTTPException(
        408,
        "Timed out waiting for Gmail authorization. "
        "Approve in the browser first, then call this endpoint again.",
    )


class SendReportRequest(BaseModel):
    run_id:  str
    to:      str
    subject: Optional[str] = None


@router.post("/send-report", summary="Email an audit report as Excel attachment")
async def send_audit_report(
    body:   SendReportRequest,
    claims: dict         = Depends(get_current_claims),
    db:     AsyncSession = Depends(get_db),
) -> dict:
    """Send a saved audit run as a formatted Excel report via Gmail.

    Builds a multi-sheet .xlsx (Summary, Course Grades, Missing Courses,
    Prereq Failures) and sends it to ``to`` with a plain-text summary in
    the email body.

    Requires prior Gmail authorization (``/api/gmail/authorize/start``).

    Args:
        run_id:  UUID of the saved audit run.
        to:      Recipient email address.
        subject: Optional subject override.
    """
    if not _EMAIL_RE.match(body.to):
        raise HTTPException(400, f"Invalid recipient email: {body.to!r}")

    user_id = claims["user_id"]
    token   = await _get_gmail_token(db, user_id)
    if not token:
        raise HTTPException(
            401,
            "Gmail not authorized. Call /api/gmail/authorize/start first.",
        )

    # Fetch the audit run — user must own it
    result = await db.execute(
        select(AuditRun).where(
            AuditRun.id == body.run_id,
            AuditRun.user_id == user_id,
        )
    )
    run = result.scalar_one_or_none()
    if not run:
        raise HTTPException(404, "Audit run not found or not owned by you.")

    audit_result = run.result or {}

    # Build Excel
    try:
        xlsx_bytes = _build_excel(audit_result)
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server.")

    subject  = body.subject
    if not subject:
        program  = audit_result.get("program", "")
        eligible = (audit_result.get("deficiency") or {}).get("eligible", False)
        subject  = f"NSU Audit Report — {program} — {'Eligible' if eligible else 'Not Eligible'}"

    body_text = _build_body(audit_result)
    filename  = f"audit_report_{str(run.id)[:8]}.xlsx"

    msg_id = await _send_via_gmail(token, body.to, subject, body_text, xlsx_bytes, filename)

    return {
        "ok":       True,
        "message":  f"Report sent to {body.to} successfully.",
        "gmail_id": msg_id,
        "subject":  subject,
        "filename": filename,
    }


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(result: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Summary"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    ok_fill  = PatternFill("solid", fgColor="C6EFCE")
    no_fill  = PatternFill("solid", fgColor="FFC7CE")

    def _hrow(sheet, row, values):
        for col, val in enumerate(values, 1):
            c = sheet.cell(row=row, column=col, value=val)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")

    _hrow(ws, 1, ["Field", "Value"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36

    deficiency = result.get("deficiency") or {}
    eligible   = deficiency.get("eligible", False)

    for r, (label, value) in enumerate([
        ("Program",           result.get("program", "")),
        ("CGPA",              result.get("cgpa", "")),
        ("Credits Completed", result.get("credit_completed", "")),
        ("Required Credits",  result.get("required_credits", "")),
        ("Academic Standing", result.get("academic_standing", "")),
        ("Eligible",          "YES" if eligible else "NO"),
        ("Credit Shortfall",  deficiency.get("credit_shortfall", 0)),
        ("On Probation",      "YES" if deficiency.get("probation") else "NO"),
        ("Run ID",            result.get("run_id", "")),
    ], 2):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=r, column=2, value=str(value))
        if label == "Eligible":
            cell.fill = ok_fill if eligible else no_fill

    breakdown = result.get("cgpa_breakdown") or result.get("courses") or []
    if breakdown:
        ws2 = wb.create_sheet("Course Grades")
        headers = list(breakdown[0].keys())
        _hrow(ws2, 1, headers)
        for col_i in range(1, len(headers) + 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = 18
        for r, course in enumerate(breakdown, 2):
            for c, key in enumerate(headers, 1):
                ws2.cell(row=r, column=c, value=course.get(key, ""))

    missing = deficiency.get("missing_mandatory") or []
    if missing:
        ws3 = wb.create_sheet("Missing Courses")
        _hrow(ws3, 1, ["Category", "Missing Courses"])
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 50
        for r, item in enumerate(missing, 2):
            ws3.cell(row=r, column=1, value=item.get("category", ""))
            ws3.cell(row=r, column=2, value=", ".join(item.get("courses", [])))

    prereqs = deficiency.get("prereq_failures_list") or []
    if prereqs:
        ws4 = wb.create_sheet("Prereq Failures")
        _hrow(ws4, 1, ["Course", "Reason"])
        ws4.column_dimensions["A"].width = 16
        ws4.column_dimensions["B"].width = 60
        for r, item in enumerate(prereqs, 2):
            ws4.cell(row=r, column=1, value=item.get("course", ""))
            ws4.cell(row=r, column=2, value=item.get("reason", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_body(result: dict) -> str:
    deficiency = result.get("deficiency") or {}
    eligible   = deficiency.get("eligible", False)
    lines = [
        "NSU Graduation Audit Report",
        "=" * 40,
        f"Program          : {result.get('program', 'N/A')}",
        f"CGPA             : {result.get('cgpa', 'N/A')}",
        f"Credits Earned   : {result.get('credit_completed', 'N/A')} / {result.get('required_credits', 'N/A')}",
        f"Academic Standing: {result.get('academic_standing', 'N/A')}",
        f"Eligible         : {'YES ✓' if eligible else 'NO ✗'}",
        "",
    ]
    if not eligible:
        lines += ["DEFICIENCIES", "-" * 40]
        shortfall = deficiency.get("credit_shortfall", 0)
        if shortfall and float(shortfall) > 0:
            lines.append(f"• Credit shortfall: {shortfall} cr")
        if deficiency.get("probation"):
            lines.append("• On academic probation (CGPA < 2.0)")
        for item in (deficiency.get("missing_mandatory") or []):
            lines.append(f"• Missing [{item.get('category','')}]: {', '.join(item.get('courses',[]))}")
        for item in (deficiency.get("prereq_failures_list") or []):
            lines.append(f"• Prereq failure — {item.get('course','')}: {item.get('reason','')}")
        lines.append("")
    lines += ["The full breakdown is attached as an Excel file.", "", "— NSU Audit System"]
    return "\n".join(lines)


async def _send_via_gmail(
    token: str,
    to: str,
    subject: str,
    body_text: str,
    attachment_bytes: bytes,
    filename: str,
) -> str:
    msg = email.mime.multipart.MIMEMultipart()
    msg["to"]      = to
    msg["subject"] = subject
    msg.attach(email.mime.text.MIMEText(body_text, "plain"))
    part = email.mime.application.MIMEApplication(attachment_bytes, Name=filename)
    part["Content-Disposition"] = f'attachment; filename="{filename}"'
    msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()

    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(
            _GMAIL_SEND_URL,
            headers={"Authorization": f"Bearer {token}"},
            json={"raw": raw},
        )

    if resp.status_code == 401:
        raise HTTPException(401, "Gmail token expired. Re-authorize via /api/gmail/authorize/start.")
    if resp.status_code not in (200, 201):
        raise HTTPException(502, f"Gmail send failed: {resp.status_code} — {resp.text}")

    return resp.json().get("id", "")
